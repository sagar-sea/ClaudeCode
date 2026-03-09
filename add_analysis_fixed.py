import pandas as pd
from openpyxl import load_workbook

# Load the existing workbook
wb = load_workbook('combined_names.xlsx')
ws = wb['Names']

# Read the data from the worksheet
data = []
for row in range(2, ws.max_row + 1):
    name = ws[f'A{row}'].value
    lastname = ws[f'B{row}'].value
    fullname = ws[f'C{row}'].value
    data.append({'Name': name, 'LastName': lastname, 'FullName': fullname})

# Convert to DataFrame for easier manipulation
df = pd.DataFrame(data)

# Create a new worksheet for the analysis
analysis_ws = wb.create_sheet('Analysis')

# Add headers for analysis
analysis_ws['A1'] = 'Analysis Type'
analysis_ws['B1'] = 'Value'

# Add some basic statistics
analysis_ws['A2'] = 'Total Records'
analysis_ws['B2'] = len(df)

# Count unique names
analysis_ws['A3'] = 'Unique First Names'
analysis_ws['B3'] = len(df['Name'].unique())

# Count unique last names
analysis_ws['A4'] = 'Unique Last Names'
analysis_ws['B4'] = len(df['LastName'].unique())

# Add another section for name frequencies
analysis_ws['A6'] = 'First Name Frequencies'
analysis_ws['A7'] = 'Name'
analysis_ws['B7'] = 'Count'

# Calculate frequency of first names
name_counts = df['Name'].value_counts()
row_num = 8
for name, count in name_counts.items():
    analysis_ws[f'A{row_num}'] = name
    analysis_ws[f'B{row_num}'] = count
    row_num += 1

# Add another section for last name frequencies
next_row = row_num + 2
analysis_ws[f'A{next_row}'] = 'Last Name Frequencies'
analysis_ws[f'A{next_row + 1}'] = 'LastName'
analysis_ws[f'B{next_row + 1}'] = 'Count'

# Calculate frequency of last names
lastname_counts = df['LastName'].value_counts()
row_num = next_row + 3
for lastname, count in lastname_counts.items():
    analysis_ws[f'A{row_num}'] = lastname
    analysis_ws[f'B{row_num}'] = count
    row_num += 1

# Save the workbook
wb.save('combined_names.xlsx')
print("Analysis sheet added to 'combined_names.xlsx'!")