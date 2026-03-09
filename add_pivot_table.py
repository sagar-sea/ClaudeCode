import pandas as pd
from openpyxl import load_workbook
from openpyxl.pivot.table import PivotTable, PivotField
from openpyxl.pivot.fields import PivotArea
from openpyxl.pivot.record import Record
from openpyxl.worksheet.table import Table, TableStyleInfo

# Load the existing workbook
wb = load_workbook('combined_names.xlsx')
ws = wb['Names']

# Add a new worksheet for the pivot table
pivot_ws = wb.create_sheet('PivotTable')

# Create the pivot table
# For this example, we'll create a simple pivot table that shows counts
# Since all our data is text, we'll create a pivot table that shows the count of each combination

# Write headers for pivot table
pivot_ws['A1'] = 'Name'
pivot_ws['B1'] = 'Count'

# Since we're working with text data, let's create a simple frequency count
# We'll count how many times each first name appears
data = []
for row in range(2, ws.max_row + 1):
    name = ws[f'A{row}'].value
    lastname = ws[f'B{row}'].value
    fullname = ws[f'C{row}'].value
    data.append((name, lastname, fullname))

# Count occurrences of each name
name_counts = {}
for name, lastname, fullname in data:
    if name in name_counts:
        name_counts[name] += 1
    else:
        name_counts[name] = 1

# Write the pivot data
row_num = 2
for name, count in name_counts.items():
    pivot_ws[f'A{row_num}'] = name
    pivot_ws[f'B{row_num}'] = count
    row_num += 1

# Save the workbook
wb.save('combined_names.xlsx')
print("Pivot table added to 'combined_names.xlsx' in a new sheet called 'PivotTable'!")