from openpyxl import Workbook
from openpyxl.styles import Font

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Names"

# Add headers
ws['A1'] = 'Name'
ws['B1'] = 'LastName'
ws['C1'] = 'FullName'

# Add some sample data
sample_data = [
    ['John', 'Doe'],
    ['Jane', 'Smith'],
    ['Michael', 'Johnson'],
    ['Sarah', 'Williams'],
    ['David', 'Brown']
]

# Populate the data
for i, (first, last) in enumerate(sample_data, start=2):
    ws[f'A{i}'] = first
    ws[f'B{i}'] = last
    ws[f'C{i}'] = f'=A{i}&" "&B{i}'  # Formula to combine first and last names

# Format headers
for cell in ws[1]:
    cell.font = Font(bold=True)

# Adjust column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 25

# Save the workbook
wb.save('combined_names.xlsx')
print("Excel file 'combined_names.xlsx' created successfully!")